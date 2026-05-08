"""Run the Golden Questions through a deployed /api/ask endpoint.

For each question, scores:
  - recall@5, recall@10: fraction of expected_key_resource_ids that appeared in retrieved chunks
  - mrr: mean reciprocal rank (1/rank of first expected doc in retrieved list)
  - did_answer: did the system produce a non-fallback answer?
  - latency: end-to-end time
  - generated_answer: the actual response (for human review)

Summary includes: mean recall@5, recall@10, MRR, latency p50/p95, per-phase breakdown.

Run:
    python -m eval.run_golden \\
        --workbook data/metadata/Enable_Stocktake_v6.xlsx \\
        --endpoint http://localhost:3000/api/ask \\
        --out data/derived/eval_results.csv

Spot-check mode (print full answers alongside expected docs):
    python -m eval.run_golden ... --spot-check
    python -m eval.run_golden ... --spot-check --phase "Assessment & Planning"
"""
from __future__ import annotations

import argparse
import csv
import json
import statistics
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from tqdm import tqdm


def parse_expected_ids(s: str) -> list[str]:
    """Parse 'ks_087; ks_068; ks_058' → ['ks_087', 'ks_068', 'ks_058']."""
    if not s or "(none" in s:
        return []
    return [x.strip() for x in s.replace(",", ";").split(";") if x.strip().startswith("ks_")]


def load_questions(workbook: Path) -> list[dict]:
    """Load Golden Questions from the workbook."""
    wb = load_workbook(workbook, data_only=True)
    if "Golden_Questions" not in wb.sheetnames:
        raise RuntimeError(f"Workbook has no Golden_Questions sheet: {workbook}")
    ws = wb["Golden_Questions"]

    # Find the header row (skip title row if present)
    header_row = None
    for r in range(1, min(5, ws.max_row + 1)):
        if ws.cell(r, 1).value == "#":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("Couldn't locate header row in Golden_Questions")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    h_idx = {h: i for i, h in enumerate(headers) if h}

    questions = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not row[0]:
            continue
        questions.append({
            "n": row[h_idx["#"]],
            "question": row[h_idx["Question"]],
            "phase": row[h_idx["Phase"]],
            "mode": (row[h_idx["Mode"]] or "Answer").lower(),
            "expected_doc_types": row[h_idx.get("Expected doc types", -1)] if "Expected doc types" in h_idx else "",
            "expected_ids": parse_expected_ids(row[h_idx["Expected key resource_ids"]] or ""),
        })
    return questions


def call_ask(endpoint: str, question: str, mode: str) -> tuple[dict, float]:
    """Call /api/ask. Returns (parsed_response, latency_seconds)."""
    t0 = time.time()
    resp = requests.post(
        endpoint,
        json={"question": question, "mode": mode if mode in ("answer", "cite") else "answer"},
        timeout=60,
    )
    elapsed = time.time() - t0
    resp.raise_for_status()
    return resp.json(), elapsed


def _reciprocal_rank(expected: set[str], retrieved: list[str]) -> float:
    """1 / rank of first expected doc in retrieved list. 0 if not found."""
    for i, rid in enumerate(retrieved, 1):
        if rid in expected:
            return 1.0 / i
    return 0.0


def score_question(q: dict, response: dict) -> dict:
    """Compute eval metrics for one question."""
    expected = set(q["expected_ids"])
    retrieved_ids = [c.get("resource_id") for c in response.get("citations", [])]
    retrieved_unique: list[str] = []
    seen: set[str] = set()
    for r in retrieved_ids:
        if r and r not in seen:
            retrieved_unique.append(r)
            seen.add(r)

    def recall_at(k: int) -> float:
        if not expected:
            return -1.0  # N/A
        top = set(retrieved_unique[:k])
        hits = expected & top
        return len(hits) / len(expected)

    mrr = _reciprocal_rank(expected, retrieved_unique) if expected else -1.0

    answer_text = response.get("answer", "") or ""
    is_fallback = any(phrase in answer_text.lower() for phrase in [
        "not in the indexed corpus",
        "not in the corpus",
        "i don't have information",
        "no relevant information",
    ])

    return {
        "recall_at_5": recall_at(5),
        "recall_at_10": recall_at(10),
        "mrr": mrr,
        "did_answer": not is_fallback,
        "n_retrieved": len(retrieved_unique),
        "retrieved_top_5": "; ".join(retrieved_unique[:5]),
        "answer_chars": len(answer_text),
        "answer_preview": answer_text[:200],
        "answer_full": answer_text,
    }


def _percentile(values: list[float], p: float) -> float:
    """p-th percentile (0–100) of a sorted-or-unsorted list."""
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    idx = (p / 100) * (len(sorted_vals) - 1)
    lo, hi = int(idx), min(int(idx) + 1, len(sorted_vals) - 1)
    frac = idx - lo
    return sorted_vals[lo] + frac * (sorted_vals[hi] - sorted_vals[lo])


def print_summary(results: list[dict]) -> None:
    valid = [r for r in results if r["recall_at_5"] >= 0]
    if not valid:
        print("No answerable questions scored.")
        return

    mean_r5 = sum(r["recall_at_5"] for r in valid) / len(valid)
    mean_r10 = sum(r["recall_at_10"] for r in valid) / len(valid)
    mean_mrr = sum(r["mrr"] for r in valid if r["mrr"] >= 0) / max(
        1, sum(1 for r in valid if r["mrr"] >= 0)
    )

    latencies = [r["latency_sec"] for r in results if r["latency_sec"] > 0]
    p50 = _percentile(latencies, 50)
    p95 = _percentile(latencies, 95)

    print(f"\nSummary across {len(valid)} answerable questions ({len(results)} total):")
    print(f"  Mean recall@5:  {mean_r5:.3f}")
    print(f"  Mean recall@10: {mean_r10:.3f}")
    print(f"  Mean MRR:       {mean_mrr:.3f}")
    print(f"  Latency p50:    {p50:.1f}s")
    print(f"  Latency p95:    {p95:.1f}s")

    # Per-phase breakdown
    phases = sorted({r["phase"] for r in valid if r["phase"]})
    if phases:
        print(f"\n  {'Phase':<22} {'N':>3}  {'R@5':>5}  {'R@10':>5}  {'MRR':>5}")
        print("  " + "-" * 50)
        for phase in phases:
            subset = [r for r in valid if r["phase"] == phase]
            r5 = sum(r["recall_at_5"] for r in subset) / len(subset)
            r10 = sum(r["recall_at_10"] for r in subset) / len(subset)
            phase_mrr_vals = [r["mrr"] for r in subset if r["mrr"] >= 0]
            pmrr = sum(phase_mrr_vals) / len(phase_mrr_vals) if phase_mrr_vals else 0.0
            print(f"  {phase:<22} {len(subset):>3}  {r5:>5.2f}  {r10:>5.2f}  {pmrr:>5.2f}")


def print_spot_check(q: dict, result: dict) -> None:
    sep = "─" * 72
    print(f"\n{sep}")
    print(f"Q{q['n']}  [{q['phase']}]  mode={q['mode']}")
    print(f"QUESTION: {q['question']}")
    print(f"\nEXPECTED DOCS: {', '.join(q['expected_ids']) or '(none)'}")
    print(f"RETRIEVED TOP-5: {result.get('retrieved_top_5', '')}")
    print(f"recall@5={result['recall_at_5']:.2f}  MRR={result['mrr']:.2f}  latency={result['latency_sec']:.1f}s")
    print(f"\nGENERATED ANSWER:\n{result.get('answer_full', '')}")
    print(sep)


def run_eval(
    workbook: Path,
    endpoint: str,
    out_csv: Path,
    spot_check: bool = False,
    phase_filter: str | None = None,
) -> dict:
    """Run all Golden Questions and write a results CSV."""
    questions = load_questions(workbook)
    if phase_filter:
        questions = [q for q in questions if q["phase"] == phase_filter]
    print(f"Loaded {len(questions)} Golden Questions from {workbook.name}")
    print(f"Endpoint: {endpoint}\n")

    results = []
    for q in tqdm(questions, desc="Running eval"):
        try:
            response, latency = call_ask(endpoint, q["question"], q["mode"])
            metrics = score_question(q, response)
            metrics.update({
                "n": q["n"],
                "phase": q["phase"],
                "mode": q["mode"],
                "question": q["question"][:120],
                "expected_ids": "; ".join(q["expected_ids"]),
                "latency_sec": round(latency, 2),
                "error": "",
            })
        except Exception as e:
            metrics = {
                "n": q["n"], "phase": q["phase"], "mode": q["mode"],
                "question": q["question"][:120],
                "expected_ids": "; ".join(q["expected_ids"]),
                "latency_sec": 0,
                "recall_at_5": -1, "recall_at_10": -1, "mrr": -1,
                "did_answer": False, "n_retrieved": 0,
                "retrieved_top_5": "", "answer_chars": 0,
                "answer_preview": "", "answer_full": "", "error": str(e)[:200],
            }
        results.append(metrics)

        if spot_check:
            print_spot_check(q, metrics)

    print_summary(results)

    # Write CSV (answer_full excluded — too wide)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "n", "phase", "mode", "question", "expected_ids",
        "recall_at_5", "recall_at_10", "mrr",
        "did_answer", "n_retrieved", "retrieved_top_5",
        "latency_sec", "answer_chars", "answer_preview", "error",
    ]
    with out_csv.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in results:
            w.writerow(r)
    print(f"\nWrote results: {out_csv}")
    return {"n_questions": len(results), "n_answerable": sum(1 for r in results if r["recall_at_5"] >= 0)}


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--endpoint", required=True, type=str)
    parser.add_argument("--out", default=Path("data/derived/eval_results.csv"), type=Path)
    parser.add_argument(
        "--spot-check",
        action="store_true",
        help="Print the full generated answer and expected docs after each question.",
    )
    parser.add_argument(
        "--phase",
        default=None,
        type=str,
        help="Filter to a single phase (exact match), e.g. 'Assessment & Planning'.",
    )
    args = parser.parse_args()
    run_eval(
        args.workbook,
        args.endpoint,
        args.out,
        spot_check=args.spot_check,
        phase_filter=args.phase,
    )
