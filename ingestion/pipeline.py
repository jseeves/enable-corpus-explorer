"""End-to-end ingestion orchestrator.

Steps:
  1. Extract text from each PDF in --pdf-dir, matched against the workbook by file_name
  2. (Optional) Enrich metadata via Claude — produces a v7 workbook
  3. Chunk the extracted documents using the (possibly enriched) workbook metadata
  4. Embed chunks with Voyage and upsert to Pinecone
  5. Compute UMAP coords for the explorer

Run:
    python -m ingestion.pipeline \\
        --metadata data/metadata/Enable_Stocktake_v6.xlsx \\
        --pdf-dir data/raw_pdfs \\
        --enrich

Add --skip-embed if you only want to refresh extraction/chunking without re-upserting.
Add --incremental to skip docs whose resource_id is already in the index.
"""
from __future__ import annotations

import argparse
import json
import os
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from tqdm import tqdm

from ingestion import chunk as chunk_mod
from ingestion import build_visualization
from ingestion import embed_index
from ingestion import enrich as enrich_mod
from ingestion.extract import ExtractedDocument, extract_one


def step_extract(workbook: Path, pdf_dir: Path, out_jsonl: Path) -> int:
    """Step 1: extract text from each PDF in the workbook."""
    wb = load_workbook(workbook, data_only=True)
    ws = wb["Corpus_Classification"]
    headers = [c.value for c in ws[1]]
    h_idx = {h: i + 1 for i, h in enumerate(headers)}

    docs: list[ExtractedDocument] = []
    for r in tqdm(range(2, ws.max_row + 1), desc="Extracting"):
        rid = ws.cell(r, 1).value
        fname = ws.cell(r, h_idx["file_name"]).value
        if not rid or not fname:
            continue
        # Skip duplicates
        short = ws.cell(r, h_idx["short_summary"]).value or ""
        if short.startswith("DUPLICATE"):
            continue
        path = pdf_dir / fname
        if not path.exists():
            print(f"  ⚠ missing: {fname}")
            continue
        try:
            docs.append(extract_one(path, rid))
        except Exception as e:
            print(f"  ⚠ extraction failed for {rid}: {e}")

    out_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with out_jsonl.open("w") as f:
        for d in docs:
            f.write(json.dumps(d.to_dict()) + "\n")

    return len(docs)


def main() -> None:
    load_dotenv()

    parser = argparse.ArgumentParser()
    parser.add_argument("--metadata", required=True, type=Path,
                        help="Path to Enable_Stocktake_v6.xlsx (or your team's workbook)")
    parser.add_argument("--pdf-dir", required=True, type=Path,
                        help="Directory containing the corpus PDFs")
    parser.add_argument("--enrich", action="store_true",
                        help="Run Track 2 metadata enrichment via Claude (produces v7 workbook)")
    parser.add_argument("--skip-embed", action="store_true",
                        help="Skip step 4 (embedding + Pinecone upsert)")
    parser.add_argument("--skip-umap", action="store_true",
                        help="Skip step 5 (UMAP coordinates)")
    parser.add_argument("--incremental", action="store_true",
                        help="Only ingest docs whose resource_id isn't already in the index")
    parser.add_argument("--derived", default=Path("data/derived"), type=Path,
                        help="Directory for intermediate outputs")
    args = parser.parse_args()

    args.derived.mkdir(parents=True, exist_ok=True)
    extracted_jsonl = args.derived / "extracted.jsonl"
    chunks_jsonl = args.derived / "chunks.jsonl"
    enriched_workbook = args.derived / "enriched_metadata_v7.xlsx"
    umap_json = args.derived / "umap.json"

    # ---- Step 1: extract ----
    print("\n[1/5] Extracting text from PDFs...")
    n = step_extract(args.metadata, args.pdf_dir, extracted_jsonl)
    print(f"      → {n} docs extracted to {extracted_jsonl}")

    # ---- Step 2: enrich (optional) ----
    metadata_for_chunks = args.metadata
    if args.enrich:
        print("\n[2/5] Enriching metadata via Claude (Track 2)...")
        stats = enrich_mod.enrich_corpus(args.metadata, args.pdf_dir, enriched_workbook)
        print(f"      → enriched {stats['enriched']}/{stats['processed']} docs")
        print(f"      → enriched workbook: {enriched_workbook}")
        metadata_for_chunks = enriched_workbook
    else:
        print("\n[2/5] Skipping enrichment (use --enrich to run Track 2)")

    # ---- Step 3: chunk ----
    print("\n[3/5] Chunking documents...")
    stats = chunk_mod.chunk_corpus(extracted_jsonl, metadata_for_chunks, chunks_jsonl)
    print(f"      → {stats['n_docs']} docs → {stats['n_chunks']} chunks")

    # ---- Step 4: embed + upsert ----
    if not args.skip_embed:
        print("\n[4/5] Embedding chunks (Voyage) and upserting to Pinecone...")
        upsert_stats = embed_index.upsert_chunks(chunks_jsonl)
        print(f"      → upserted {upsert_stats['n_upserted']} vectors to '{upsert_stats['index']}'")
    else:
        print("\n[4/5] Skipping embed/upsert (--skip-embed)")

    # ---- Step 5: UMAP ----
    if not args.skip_umap:
        print("\n[5/5] Computing UMAP coordinates for the explorer...")
        build_visualization.build_corpus_json(metadata_for_chunks, chunks_jsonl, umap_json)
    else:
        print("\n[5/5] Skipping UMAP (--skip-umap)")

    print(f"\n✓ Pipeline complete. Outputs in {args.derived}/")
    print(f"   - extracted.jsonl")
    print(f"   - chunks.jsonl")
    if args.enrich:
        print(f"   - enriched_metadata_v7.xlsx")
    if not args.skip_umap:
        print(f"   - umap.json (copy this to app/public/ for the explorer)")
    print(f"\nNext: cd app && pnpm dev")


if __name__ == "__main__":
    main()
