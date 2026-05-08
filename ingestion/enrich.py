"""Track 2: per-doc metadata enrichment via Claude.

For each document, send a representative sample of the full text plus the current
metadata row, and ask Claude to produce a refined version of:
  - long_summary (300-500 words, grounded in full doc text)
  - thematic_tags (full vocabulary scan against the controlled list)
  - programs_referenced, frameworks_referenced
  - countries_covered (all countries actually discussed)
  - key_metrics_present (vocabulary-aligned, evidence-only)
  - data_vintage (with evidence quote)
  - funder (extracted from acknowledgments)

Output: an enriched copy of the workbook (data/derived/enriched_metadata.xlsx)
with diffs noted in access_notes for transparency.

This is the single biggest quality lift available — it makes long_summary,
thematic_tags, etc. reflect the whole document rather than the front pages.

Run with: python -m ingestion.enrich --workbook ... --pdf-dir ... --out ...
Cost: roughly $1-2 for a 100-doc corpus using Haiku, $5-10 with Sonnet.
"""
from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from anthropic import Anthropic
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from tqdm import tqdm

from ingestion.extract import extract_one


ENRICH_SYSTEM = """You are a metadata enrichment agent for a restoration knowledge corpus.
You read a document and produce refined structured metadata grounded exclusively in the document text.

Per-field guidance:

long_summary (300–500 words):
  Cover: core argument or purpose, methodological approach, key evidence and data sources,
  geographic scope, and main conclusions or recommendations. Write in present tense, third person.
  Do NOT copy-paste the abstract. Synthesize from the full document, not just the front pages.

thematic_tags:
  Scan the ENTIRE document for themes, not just the abstract. A report on forest carbon may also
  deserve tags like "private-sector finance" or "smallholder livelihoods" if those topics appear
  substantively (more than passing mention). Match exact vocabulary values.

programs_referenced:
  Include only programs explicitly named in the text (e.g., "AFR100", "20x20", "Bonn Challenge").
  Do not infer from geography. Exact names only.

frameworks_referenced:
  Include only frameworks explicitly cited or described. Common ones: Bonn Challenge, AFR100,
  UN Decade on Ecosystem Restoration, Paris Agreement / NDCs, UNCCD Land Degradation Neutrality,
  TNFD, SBTN. Exact names only.

countries_covered:
  List every country discussed substantively — not just named in passing. Check body text, tables,
  and case studies, not just the title or introduction. A "global" analysis may have deep case
  studies in 10 countries; list them all. Use standard country names.

key_metrics_present:
  Only flag metrics you can find stated in the document with actual numbers. If the document
  discusses costs conceptually but gives no figures, do NOT tag "cost_per_ha". Evidence required.

data_vintage:
  State the date range AND quote the specific evidence. Format: "YYYY–YYYY; [quote and location]".
  Example: "2010–2019; Table 2 caption: 'monitoring data collected 2010–2019 across 47 sites'".
  If multiple date ranges apply, list all. If not determinable, use "unknown".

funder:
  Extract from the acknowledgments or funding statement section. If not present, use "unknown".
  Multiple funders: semicolon-separated.

enrichment_notes:
  One to three sentences describing what changed from the current metadata and why. Be specific:
  "Expanded countries_covered from ['Brazil'] to ['Brazil', 'Colombia', 'Peru', 'Mexico'] based
  on case studies in Section 3. Revised data_vintage from 2015 to 2008–2015 based on Table 1."

Rules:
- Every claim must be grounded in the document text. If you can't find evidence, omit the field.
- Use exact values from the controlled vocabularies provided.
- For multi-valued fields, return a JSON array.
- data_vintage MUST include an evidence quote — never just a year without provenance.
- Output a single JSON object. No prose, no markdown fences, no commentary before or after.
"""


def sample_doc_text(doc_text: str, head: int = 20000, tail: int = 10000) -> str:
    """Return a representative sample for long documents.

    Strategy: keep the first `head` characters (intro, exec summary, methodology)
    and the last `tail` characters (conclusions, acknowledgments, references header).
    For very long documents a middle slice is also included.
    Total budget: ~35K characters, well within Claude's context.
    """
    total = head + tail
    if len(doc_text) <= total:
        return doc_text

    head_text = doc_text[:head]
    tail_text = doc_text[-tail:]

    # Add a middle slice for very long documents (>100K chars)
    if len(doc_text) > 100_000:
        mid_start = len(doc_text) // 2 - 2500
        mid_text = doc_text[mid_start: mid_start + 5000]
        return (
            head_text
            + f"\n\n[... middle excerpt (~{mid_start // 1000}k chars in) ...]\n\n"
            + mid_text
            + "\n\n[... document truncated, showing final section ...]\n\n"
            + tail_text
        )

    return (
        head_text
        + "\n\n[... document truncated for length, showing final section ...]\n\n"
        + tail_text
    )


def build_user_prompt(doc_text: str, current_metadata: dict, vocabularies: dict) -> str:
    """Build the user message for the enrichment call."""
    sampled = sample_doc_text(doc_text)

    return f"""Refine the metadata for this document.

CURRENT METADATA (some fields may be incomplete or inaccurate — improve where you have evidence):
{json.dumps(current_metadata, indent=2)}

CONTROLLED VOCABULARIES (use these exact values where applicable):
{json.dumps(vocabularies, indent=2)}

DOCUMENT TEXT:
{sampled}

Return a JSON object with ALL of the following keys. Omit a key only if you have no grounded
evidence to populate it at all:
- long_summary: string, 300–500 words, synthesized from full document content
- thematic_tags: array of vocabulary values from Vocab_thematic_tags
- programs_referenced: array of vocabulary values from Vocab_programs_referenced
- frameworks_referenced: array of vocabulary values from Vocab_frameworks_referenced
- countries_covered: array of country names actually discussed substantively
- key_metrics_present: array of vocabulary values from Vocab_key_metrics_present (evidence-only)
- data_vintage: string with evidence quote, e.g. "2010–2018; p.12: 'data from 2010 to 2018'"
- funder: string, semicolon-separated funders from acknowledgments, or "unknown"
- enrichment_notes: string, 1–3 sentences on what changed and why

Return ONLY the JSON object. No markdown fences."""


def load_vocabularies(workbook_path: Path) -> dict:
    """Load controlled vocabularies from the workbook for prompt context."""
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Controlled_Vocabularies"]
    vocabs: dict[str, list[str]] = {}
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            if row[0] != current:
                current = row[0]
                vocabs[current] = []
        if row[1] and current:
            vocabs[current].append(str(row[1]))
    return vocabs


def enrich_document(doc_text: str, current_metadata: dict, vocabularies: dict,
                    client: Anthropic, model: str) -> dict:
    """Call Claude to refine one doc's metadata. Returns dict of refined fields."""
    response = client.messages.create(
        model=model,
        max_tokens=2048,
        system=ENRICH_SYSTEM,
        messages=[{"role": "user", "content": build_user_prompt(doc_text, current_metadata, vocabularies)}],
    )
    text = response.content[0].text.strip()
    # Strip markdown fences if Claude disobeys the instruction
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"  ⚠ JSON parse failed: {e}")
        print(f"  Raw response: {text[:500]}")
        return {}


def enrich_corpus(workbook_in: Path, pdf_dir: Path, workbook_out: Path,
                  model: str | None = None, dry_run: bool = False) -> dict:
    """Enrich every classified doc in the workbook. Writes a new workbook.

    Args:
        workbook_in: input metadata workbook (e.g. v6)
        pdf_dir: directory of source PDFs
        workbook_out: output workbook path (e.g. v7)
        model: Anthropic model to use (defaults to CLAUDE_ENRICH_MODEL env var, or Haiku)
        dry_run: if True, don't actually call the API; just check inputs
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key and not dry_run:
        raise RuntimeError("ANTHROPIC_API_KEY not set")

    model = model or os.environ.get("CLAUDE_ENRICH_MODEL", "claude-haiku-4-5-20251001")
    client = Anthropic(api_key=api_key) if not dry_run else None

    vocabularies = load_vocabularies(workbook_in)

    wb = load_workbook(workbook_in)
    ws = wb["Corpus_Classification"]
    headers = [c.value for c in ws[1]]
    h_idx = {h: i + 1 for i, h in enumerate(headers)}

    notes_col = h_idx["access_notes"]
    TNR = Font(name="Times New Roman", size=11)
    WRAP = Alignment(wrap_text=True, vertical="top")

    rows_processed = 0
    rows_enriched = 0
    rows_skipped = 0

    for r in tqdm(range(2, ws.max_row + 1), desc="Enriching"):
        rid = ws.cell(r, 1).value
        fname = ws.cell(r, 2).value
        if not rid or not fname:
            continue

        # Skip flagged duplicates
        short = ws.cell(r, h_idx["short_summary"]).value or ""
        if short.startswith("DUPLICATE"):
            rows_skipped += 1
            continue

        pdf_path = pdf_dir / fname
        if not pdf_path.exists():
            print(f"  ⚠ missing PDF for {rid}: {fname}")
            rows_skipped += 1
            continue

        rows_processed += 1
        if dry_run:
            continue

        # Extract text
        try:
            extracted = extract_one(pdf_path, rid)
            doc_text = "\n".join(p["text"] for p in extracted.pages)
        except Exception as e:
            print(f"  ⚠ extraction failed for {rid}: {e}")
            continue

        # Build current metadata dict
        current = {h: ws.cell(r, h_idx[h]).value for h in headers}
        current = {k: v for k, v in current.items() if v is not None}

        # Call Claude
        try:
            refined = enrich_document(doc_text, current, vocabularies, client, model)
        except Exception as e:
            print(f"  ⚠ enrichment failed for {rid}: {e}")
            continue

        if not refined:
            continue

        # Apply refinements (only fields that came back populated)
        FIELD_MAP = {
            "long_summary": "long_summary",
            "thematic_tags": "thematic_tags",
            "programs_referenced": "programs_referenced",
            "frameworks_referenced": "frameworks_referenced",
            "countries_covered": "countries_covered",
            "key_metrics_present": "key_metrics_present",
            "data_vintage": "data_vintage",
            "funder": "funder",
        }
        for src_key, dst_key in FIELD_MAP.items():
            if src_key in refined and refined[src_key]:
                val = refined[src_key]
                if isinstance(val, list):
                    val = " | ".join(str(x) for x in val)
                ws.cell(r, h_idx[dst_key], value=val).font = TNR
                ws.cell(r, h_idx[dst_key]).alignment = WRAP

        # Append enrichment note
        if "enrichment_notes" in refined and refined["enrichment_notes"]:
            cur_notes = ws.cell(r, notes_col).value or ""
            new_notes = (cur_notes + " | " if cur_notes else "") + f"v7 enrich: {refined['enrichment_notes']}"
            ws.cell(r, notes_col, value=new_notes).font = TNR
            ws.cell(r, notes_col).alignment = WRAP

        rows_enriched += 1

    if not dry_run:
        wb.save(workbook_out)

    return {"processed": rows_processed, "enriched": rows_enriched, "skipped": rows_skipped}


if __name__ == "__main__":
    import argparse
    from dotenv import load_dotenv
    load_dotenv()

    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--pdf-dir", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--model", default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    stats = enrich_corpus(args.workbook, args.pdf_dir, args.out,
                          model=args.model, dry_run=args.dry_run)
    print(f"\nProcessed: {stats['processed']}, Enriched: {stats['enriched']}, Skipped: {stats['skipped']}")
    if not args.dry_run:
        print(f"Saved: {args.out}")
