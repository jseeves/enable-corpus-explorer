"""Chunk extracted documents into ~400-character pieces with metadata propagation.

Each output chunk carries:
  - chunk_id (resource_id + position)
  - resource_id
  - text
  - page_num (the page this chunk starts on; we record start page only)
  - char_start, char_end (offsets within the doc's full text)
  - metadata (document-level: title, document_type, phase, etc.)

The metadata is propagated from the workbook so retrieval can filter by it.

Strategy:
  - Default: 400 characters per chunk, 80-character overlap.
  - Boundary-aware: prefer to break on paragraph (\\n\\n), then sentence (. ), then word.
  - Cross-page chunks: a chunk that spans a page boundary keeps the start page;
    page transitions inside the chunk are noted in the chunk's text via marker.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path

CHUNK_SIZE = 400
CHUNK_OVERLAP = 80


@dataclass
class Chunk:
    chunk_id: str
    resource_id: str
    text: str
    page_num: int
    char_start: int
    char_end: int
    metadata: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "chunk_id": self.chunk_id,
            "resource_id": self.resource_id,
            "text": self.text,
            "page_num": self.page_num,
            "char_start": self.char_start,
            "char_end": self.char_end,
            "metadata": self.metadata,
        }


def _find_break(text: str, target: int, min_pos: int) -> int:
    """Find a good break point in text near target position.

    Prefer (in order): paragraph break, sentence end, whitespace.
    Returns the position to break at (exclusive)."""
    if target >= len(text):
        return len(text)

    window_start = max(min_pos, target - 80)
    window_end = min(len(text), target + 80)
    window = text[window_start:window_end]

    # Paragraph break
    pp = window.rfind("\n\n", 0, target - window_start)
    if pp > 0:
        return window_start + pp + 2

    # Sentence end
    for delim in [". ", ".\n", "? ", "! "]:
        pos = window.rfind(delim, 0, target - window_start)
        if pos > 0:
            return window_start + pos + len(delim)

    # Whitespace
    pos = window.rfind(" ", 0, target - window_start)
    if pos > 0:
        return window_start + pos + 1

    # Hard break
    return target


def chunk_document(extracted: dict, doc_metadata: dict) -> list[Chunk]:
    """Chunk one extracted document.

    Args:
        extracted: dict with resource_id, file_name, pages (list of {page_num, text})
        doc_metadata: dict from the workbook row (title, document_type, etc.)
    """
    resource_id = extracted["resource_id"]

    # Build a single string with page boundary markers we can locate later.
    # Using a sentinel that won't appear naturally.
    SENTINEL = "\n\x1f__PAGE_{}__\x1f\n"
    parts = []
    for page in extracted["pages"]:
        parts.append(SENTINEL.format(page["page_num"]))
        parts.append(page["text"])
    full_text = "".join(parts)

    # Build a (char_offset → page_num) map by scanning sentinels.
    page_at = []  # list of (char_offset, page_num)
    sentinel_re = re.compile(r"\x1f__PAGE_(\d+)__\x1f")
    for m in sentinel_re.finditer(full_text):
        page_at.append((m.start(), int(m.group(1))))

    def page_for_offset(offset: int) -> int:
        page = 1
        for off, pg in page_at:
            if off <= offset:
                page = pg
            else:
                break
        return page

    # Strip sentinels for the actual text we chunk over.
    clean_text = sentinel_re.sub("", full_text)

    # Recompute offset map for clean_text by tracking the cumulative shift.
    # For each clean offset, derive the corresponding original offset.
    # Simpler approach: track which page each clean-text position falls in
    # by walking through the original text and clean text in parallel.
    clean_to_page = {}  # clean_offset → page
    orig_pos = 0
    clean_pos = 0
    while orig_pos < len(full_text):
        m = sentinel_re.match(full_text, orig_pos)
        if m:
            orig_pos = m.end()
            continue
        # Find current page
        clean_to_page[clean_pos] = page_for_offset(orig_pos)
        orig_pos += 1
        clean_pos += 1

    def page_for_clean_offset(offset: int) -> int:
        # Find the largest key ≤ offset
        if offset in clean_to_page:
            return clean_to_page[offset]
        # Binary search would be faster; for our scale this is fine
        keys = sorted(k for k in clean_to_page if k <= offset)
        return clean_to_page[keys[-1]] if keys else 1

    # Now chunk clean_text.
    chunks = []
    pos = 0
    chunk_idx = 0
    while pos < len(clean_text):
        target = pos + CHUNK_SIZE
        end = _find_break(clean_text, target, pos + CHUNK_SIZE // 2)
        chunk_text = clean_text[pos:end].strip()
        if chunk_text:
            page_num = page_for_clean_offset(pos)
            chunk = Chunk(
                chunk_id=f"{resource_id}_c{chunk_idx:04d}",
                resource_id=resource_id,
                text=chunk_text,
                page_num=page_num,
                char_start=pos,
                char_end=end,
                metadata=doc_metadata,
            )
            chunks.append(chunk)
            chunk_idx += 1
        # Advance with overlap
        pos = max(end - CHUNK_OVERLAP, pos + 1)
        if end >= len(clean_text):
            break

    return chunks


def chunk_corpus(extracted_jsonl: Path, metadata_workbook: Path, out_path: Path) -> dict:
    """Chunk every extracted doc and write to JSONL.

    Returns summary stats."""
    from openpyxl import load_workbook

    # Load metadata workbook
    wb = load_workbook(metadata_workbook, data_only=True)
    ws = wb["Corpus_Classification"]
    headers = [c.value for c in ws[1]]
    metadata_by_rid = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: v for h, v in zip(headers, row) if v is not None}
        rid = rec.get("resource_id")
        if rid:
            metadata_by_rid[rid] = rec

    out_path.parent.mkdir(parents=True, exist_ok=True)
    n_chunks = 0
    n_docs = 0
    with out_path.open("w") as out:
        with extracted_jsonl.open() as inp:
            for line in inp:
                doc = json.loads(line)
                rid = doc["resource_id"]
                meta = metadata_by_rid.get(rid)
                if not meta:
                    print(f"  ⚠ no metadata row for {rid}, skipping")
                    continue
                # Skip flagged duplicates (their summary starts with "DUPLICATE")
                short_summary = (meta.get("short_summary") or "")
                if short_summary.startswith("DUPLICATE"):
                    continue
                chunks = chunk_document(doc, _serialize_metadata(meta))
                for c in chunks:
                    out.write(json.dumps(c.to_dict()) + "\n")
                n_chunks += len(chunks)
                n_docs += 1

    return {"n_docs": n_docs, "n_chunks": n_chunks}


def _serialize_metadata(meta: dict) -> dict:
    """Pinecone metadata fields must be strings, numbers, booleans, or lists of strings.
    Convert pipe-separated lists to actual lists; coerce booleans; drop Nones."""
    out = {}
    list_fields = {"phase_of_restoration", "intended_audience_detailed", "thematic_tags",
                   "restoration_type", "evidence_type", "methodology", "key_metrics_present",
                   "countries_covered", "programs_referenced", "frameworks_referenced",
                   "target_audience", "tags"}
    bool_fields = {"makes_economic_case", "addresses_social_equity",
                   "describes_policies_or_incentives", "contains_geospatial",
                   "canonical_reference"}
    for k, v in meta.items():
        if v is None:
            continue
        if k in bool_fields:
            out[k] = str(v).strip().upper() == "TRUE"
        elif k in list_fields and isinstance(v, str):
            out[k] = [s.strip() for s in v.split("|") if s.strip()]
        elif isinstance(v, (str, int, float, bool)):
            # Truncate strings to 1000 chars (Pinecone metadata size limit)
            out[k] = v if not isinstance(v, str) else v[:1000]
    return out


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--extracted", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()
    stats = chunk_corpus(args.extracted, args.workbook, args.out)
    print(f"Chunked {stats['n_docs']} docs into {stats['n_chunks']} chunks → {args.out}")
