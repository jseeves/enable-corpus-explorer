"""Compute 2D UMAP coordinates for each document, for the corpus explorer (Tab 1).

Strategy:
  - Average the chunk embeddings per document to get a doc-level embedding.
  - Run UMAP (n_components=2) over the doc embeddings.
  - Output JSON keyed by resource_id with (x, y) plus doc-level metadata for hover/click.

The output file (data/derived/umap.json) is served by /api/corpus to the explorer.
"""
from __future__ import annotations

import json
import os
from collections import defaultdict
from pathlib import Path

import numpy as np
import voyageai
from openpyxl import load_workbook
from tqdm import tqdm

from ingestion.embed_index import VOYAGE_EMBEDDING_MODEL, get_voyage_client


def average_chunk_embeddings(chunks_jsonl: Path, voyage: voyageai.Client) -> dict[str, np.ndarray]:
    """Compute a doc-level embedding as the mean of its chunk embeddings.

    To save API calls, we re-embed only doc summaries here rather than re-fetching
    chunk vectors from Pinecone. This gives a clean, semantically-meaningful doc
    embedding for visualization (chunks would smooth out the doc's identity)."""
    # Load chunks
    by_rid: dict[str, list[str]] = defaultdict(list)
    with chunks_jsonl.open() as f:
        for line in f:
            c = json.loads(line)
            by_rid[c["resource_id"]].append(c["text"])

    # For each doc, build a representative text (concatenation of first ~5 chunks)
    rep_texts = {}
    for rid, texts in by_rid.items():
        rep = " ".join(texts[:5])[:8000]  # cap length
        rep_texts[rid] = rep

    # Embed in batches
    embeddings = {}
    rids = list(rep_texts.keys())
    for i in tqdm(range(0, len(rids), 32), desc="Embedding doc representations"):
        batch_rids = rids[i:i + 32]
        batch_texts = [rep_texts[r] for r in batch_rids]
        result = voyage.embed(texts=batch_texts, model=VOYAGE_EMBEDDING_MODEL, input_type="document")
        for r, v in zip(batch_rids, result.embeddings):
            embeddings[r] = np.array(v)

    return embeddings


def compute_umap(embeddings: dict[str, np.ndarray]) -> dict[str, tuple[float, float]]:
    """Project doc embeddings to 2D with UMAP."""
    import umap

    rids = list(embeddings.keys())
    X = np.stack([embeddings[r] for r in rids])

    n = X.shape[0]
    n_neighbors = max(min(15, n - 1), 2)

    reducer = umap.UMAP(
        n_components=2,
        n_neighbors=n_neighbors,
        min_dist=0.1,
        metric="cosine",
        random_state=42,
    )
    coords = reducer.fit_transform(X)

    return {r: (float(coords[i, 0]), float(coords[i, 1])) for i, r in enumerate(rids)}


def build_corpus_json(workbook_path: Path, chunks_jsonl: Path, out_path: Path) -> None:
    """Build the corpus JSON consumed by /api/corpus and the explorer."""
    voyage = get_voyage_client()
    embeddings = average_chunk_embeddings(chunks_jsonl, voyage)
    coords = compute_umap(embeddings)

    # Pull metadata from workbook
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Corpus_Classification"]
    headers = [c.value for c in ws[1]]
    h_idx = {h: i + 1 for i, h in enumerate(headers)}

    docs = []
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if not rid or rid not in coords:
            continue
        # Skip duplicates
        short = ws.cell(r, h_idx["short_summary"]).value or ""
        if short.startswith("DUPLICATE"):
            continue

        x, y = coords[rid]
        docs.append({
            "resource_id": rid,
            "title": _derive_title(ws.cell(r, h_idx["file_name"]).value or "", short),
            "file_name": ws.cell(r, h_idx["file_name"]).value,
            "document_type": ws.cell(r, h_idx["document_type"]).value,
            "phase_of_restoration": ws.cell(r, h_idx["phase_of_restoration"]).value,
            "target_audience": ws.cell(r, h_idx["target_audience"]).value,
            "region": ws.cell(r, h_idx["region"]).value,
            "publication_date": str(ws.cell(r, h_idx["publication_date"]).value or ""),
            "page_count": ws.cell(r, h_idx["page_count"]).value,
            "short_summary": short,
            "canonical_reference": str(ws.cell(r, h_idx["canonical_reference"]).value or ""),
            "umap_x": x,
            "umap_y": y,
        })

    out = {"docs": docs, "n_docs": len(docs)}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2))
    print(f"Wrote {len(docs)} docs with UMAP coords → {out_path}")


def _derive_title(file_name: str, summary: str) -> str:
    """Derive a clean title from a filename. Quick heuristic; can be improved."""
    name = file_name.rsplit(".", 1)[0]
    name = name.replace("_", " ").replace("-", " ").replace(" World Resources Institute", "").strip()
    return name


if __name__ == "__main__":
    import argparse
    from dotenv import load_dotenv
    load_dotenv()

    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--chunks", required=True, type=Path)
    parser.add_argument("--out", default=Path("data/derived/umap.json"), type=Path)
    args = parser.parse_args()

    build_corpus_json(args.workbook, args.chunks, args.out)
